import io
import json
import os
from copy import deepcopy
from datetime import datetime
from uuid import uuid4

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "data_store.json")
APP_PORT = int(os.environ.get("SAP_CR_MANAGER_PORT", "5055"))
APP_DEBUG = (os.environ.get("SAP_CR_MANAGER_DEBUG", "0").strip() == "1")

app.secret_key = os.environ.get("SAP_CR_MANAGER_SECRET_KEY", "sap-cr-manager-dev-secret")

STATUS_META = {
    "development": {
        "label": "Sviluppo",
        "tone": "development",
    },
    "quality": {
        "label": "Quality",
        "tone": "quality",
    },
    "production": {
        "label": "Produzione",
        "tone": "production",
    },
}

CR_TYPE_META = {
    "workbench": {
        "label": "Workbench",
    },
    "customizing": {
        "label": "Customizing",
    },
}

KANBAN_ORDER = ["development", "quality", "production"]

DEFAULT_DATA = {
    "clients": [],
}


def deep_copy_default() -> dict:
    return deepcopy(DEFAULT_DATA)


def sanitize_text(raw_value: str) -> str:
    return (raw_value or "").strip()


def _to_camel_case(raw_value: str) -> str:
    tokens: list[str] = []
    current_token: list[str] = []
    for character in sanitize_text(raw_value):
        if character.isalnum():
            current_token.append(character)
            continue
        if current_token:
            tokens.append("".join(current_token))
            current_token = []
    if current_token:
        tokens.append("".join(current_token))

    if not tokens:
        return "Progetto"

    return "".join(token[:1].upper() + token[1:] for token in tokens)


def build_project_export_filename(project_name: str, extension: str) -> str:
    safe_project_name = _to_camel_case(project_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"ListaCR_{safe_project_name}_{timestamp}.{extension}"


def new_id() -> str:
    return uuid4().hex[:12]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


_IT_MONTHS = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]


def format_it_datetime(iso_str: str) -> str:
    """Convert an ISO timestamp to Italian format: 25 Mar 2026 14:30"""
    try:
        dt = datetime.fromisoformat(iso_str)
        return f"{dt.day:02d} {_IT_MONTHS[dt.month - 1]} {dt.year} {dt.hour:02d}:{dt.minute:02d}"
    except (ValueError, TypeError, AttributeError):
        return iso_str or "-"


def now_it_datetime() -> str:
    dt = datetime.now()
    return f"{dt.day:02d} {_IT_MONTHS[dt.month - 1]} {dt.year} {dt.hour:02d}:{dt.minute:02d}"


def build_client_band_palette(client_name: str) -> dict:
    normalized_name = sanitize_text(client_name).lower() or "client"
    seed = sum((index + 1) * ord(char) for index, char in enumerate(normalized_name))
    hue = seed % 360
    return {
        "bg": f"hsl({hue} 72% 93%)",
        "border": f"hsl({hue} 48% 78%)",
        "accent": f"hsl({hue} 58% 34%)",
    }


def ensure_data_file() -> None:
    if os.path.exists(DATA_FILE):
        return
    save_data(deep_copy_default())


def normalize_status(raw_value: str) -> str:
    candidate = sanitize_text(raw_value).lower()
    if candidate in STATUS_META:
        return candidate
    return "development"


def normalize_cr_type(raw_value: str) -> str:
    candidate = sanitize_text(raw_value).lower()
    if candidate in CR_TYPE_META:
        return candidate
    return "workbench"


def normalize_release_order(raw_value, fallback: int = 9999) -> int:
    try:
        parsed = int(raw_value)
    except (TypeError, ValueError):
        return fallback
    if parsed < 1:
        return fallback
    return parsed


def sort_crs_for_execution(crs: list[dict]) -> list[dict]:
    return sorted(
        crs,
        key=lambda item: (
            normalize_release_order(item.get("release_order")),
            item.get("updated_at", ""),
            sanitize_text(item.get("cr_key")).lower(),
        ),
    )


def next_release_order(project: dict) -> int:
    orders = [normalize_release_order(cr.get("release_order"), fallback=0) for cr in project.get("crs", [])]
    return (max(orders) + 1) if orders else 1


def has_release_order_conflict(project: dict, release_order: int, exclude_cr_id: str | None = None) -> bool:
    for item in project.get("crs", []):
        if exclude_cr_id and item.get("id") == exclude_cr_id:
            continue
        if normalize_release_order(item.get("release_order")) == release_order:
            return True
    return False


def normalize_cr(raw_cr: dict) -> dict:
    created_at = sanitize_text(raw_cr.get("created_at")) or now_iso()
    updated_at = sanitize_text(raw_cr.get("updated_at")) or created_at
    return {
        "id": sanitize_text(raw_cr.get("id")) or new_id(),
        "cr_key": sanitize_text(raw_cr.get("cr_key")),
        "created_by": sanitize_text(raw_cr.get("created_by")),
        "description": sanitize_text(raw_cr.get("description")),
        "notes": sanitize_text(raw_cr.get("notes")),
        "status": normalize_status(raw_cr.get("status")),
        "cr_type": normalize_cr_type(raw_cr.get("cr_type")),
        "release_order": normalize_release_order(raw_cr.get("release_order")),
        "created_at": created_at,
        "updated_at": updated_at,
    }


def normalize_project(raw_project: dict) -> dict:
    crs = [normalize_cr(item) for item in raw_project.get("crs", []) if isinstance(item, dict)]
    return {
        "id": sanitize_text(raw_project.get("id")) or new_id(),
        "name": sanitize_text(raw_project.get("name")),
        "crs": crs,
    }


def normalize_client(raw_client: dict) -> dict:
    projects = [normalize_project(item) for item in raw_client.get("projects", []) if isinstance(item, dict)]
    return {
        "id": sanitize_text(raw_client.get("id")) or new_id(),
        "name": sanitize_text(raw_client.get("name")),
        "projects": projects,
    }


def load_data() -> dict:
    ensure_data_file()
    with open(DATA_FILE, "r", encoding="utf-8") as handle:
        loaded = json.load(handle)

    data = deep_copy_default()
    data["clients"] = [normalize_client(item) for item in loaded.get("clients", []) if isinstance(item, dict)]
    return data


def save_data(data: dict) -> None:
    with open(DATA_FILE, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)


def find_client(data: dict, client_id: str) -> dict | None:
    for client in data["clients"]:
        if client.get("id") == client_id:
            return client
    return None


def find_project(client: dict, project_id: str) -> dict | None:
    for project in client.get("projects", []):
        if project.get("id") == project_id:
            return project
    return None


def find_cr(project: dict, cr_id: str) -> dict | None:
    for cr in project.get("crs", []):
        if cr.get("id") == cr_id:
            return cr
    return None


def count_statuses(clients: list[dict]) -> dict:
    counts = {key: 0 for key in STATUS_META}
    for client in clients:
        for project in client.get("projects", []):
            for cr in project.get("crs", []):
                status = normalize_status(cr.get("status"))
                counts[status] += 1
    return counts


def matches_filters(cr: dict, search_text: str, status_filter: str, type_filter: str = "all") -> bool:
    if status_filter != "all" and normalize_status(cr.get("status")) != status_filter:
        return False

    if type_filter != "all" and normalize_cr_type(cr.get("cr_type")) != type_filter:
        return False

    if not search_text:
        return True

    haystack = " ".join(
        [
            sanitize_text(cr.get("cr_key")),
            sanitize_text(cr.get("created_by")),
            sanitize_text(cr.get("description")),
            sanitize_text(cr.get("notes")),
        ]
    ).lower()
    return search_text in haystack


def normalize_filter_value(raw_value: str) -> str:
    return sanitize_text(raw_value).lower()


def build_filter_options(data: dict) -> dict:
    unique_clients: dict[str, str] = {}
    for client in data["clients"]:
        client_name = sanitize_text(client.get("name"))
        if not client_name:
            continue
        unique_clients.setdefault(client_name.lower(), client_name)

    client_names = sorted(unique_clients.values(), key=str.lower)
    client_filters = []
    for client_name in client_names:
        palette = build_client_band_palette(client_name)
        client_filters.append(
            {
                "name": client_name,
                "filter_value": client_name.lower(),
                "tab_bg": palette["bg"],
                "tab_border": palette["border"],
                "tab_accent": palette["accent"],
            }
        )

    project_names = sorted(
        {
            sanitize_text(project.get("name"))
            for client in data["clients"]
            for project in client.get("projects", [])
            if sanitize_text(project.get("name"))
        },
        key=str.lower,
    )
    return {
        "clients": client_filters,
        "projects": project_names,
        "cr_types": [CR_TYPE_META["workbench"]["label"], CR_TYPE_META["customizing"]["label"]],
    }


def build_global_kanban_columns(global_crs: list[dict]) -> list[dict]:
    columns = []
    for status_key in KANBAN_ORDER:
        meta = STATUS_META[status_key]
        columns.append(
            {
                "key": status_key,
                "label": meta["label"],
                "tone": meta["tone"],
                "crs": [cr for cr in global_crs if cr.get("status") == status_key],
            }
        )
    return columns


def build_export_rows(
    search_text: str = "",
    status_filter: str = "all",
    client_filter: str = "all",
    project_filter: str = "all",
    type_filter: str = "all",
) -> list[dict]:
    data = load_data()
    normalized_search = sanitize_text(search_text).lower()
    normalized_status = sanitize_text(status_filter).lower() or "all"
    normalized_client = normalize_filter_value(client_filter) or "all"
    normalized_project = normalize_filter_value(project_filter) or "all"
    normalized_type = normalize_filter_value(type_filter) or "all"
    if normalized_status != "all" and normalized_status not in STATUS_META:
        normalized_status = "all"
    if normalized_type != "all" and normalized_type not in CR_TYPE_META:
        normalized_type = "all"

    rows: list[dict] = []
    sorted_clients = sorted(data["clients"], key=lambda item: item.get("name", "").lower())
    for client in sorted_clients:
        client_name = sanitize_text(client.get("name"))
        if normalized_client != "all" and client_name.lower() != normalized_client:
            continue

        sorted_projects = sorted(client.get("projects", []), key=lambda item: item.get("name", "").lower())
        for project in sorted_projects:
            project_name = sanitize_text(project.get("name"))
            if normalized_project != "all" and project_name.lower() != normalized_project:
                continue

            sorted_crs = sort_crs_for_execution(project.get("crs", []))
            for cr in sorted_crs:
                if not matches_filters(cr, normalized_search, normalized_status, normalized_type):
                    continue

                status = normalize_status(cr.get("status"))
                cr_type = normalize_cr_type(cr.get("cr_type"))
                rows.append(
                    {
                        "client": client_name,
                        "project": project_name,
                        "release_order": normalize_release_order(cr.get("release_order")),
                        "cr_type": CR_TYPE_META[cr_type]["label"],
                        "cr_key": sanitize_text(cr.get("cr_key")),
                        "created_by": sanitize_text(cr.get("created_by")),
                        "description": sanitize_text(cr.get("description")),
                        "notes": sanitize_text(cr.get("notes")),
                        "status": STATUS_META[status]["label"],
                    }
                )

    return rows


def build_project_export_rows(client_id: str, project_id: str) -> dict:
    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        raise ValueError("Cliente non trovato.")

    project = find_project(client, project_id)
    if project is None:
        raise ValueError("Progetto non trovato.")

    client_name = sanitize_text(client.get("name"))
    project_name = sanitize_text(project.get("name"))
    sorted_crs = sort_crs_for_execution(project.get("crs", []))

    rows = []
    for cr in sorted_crs:
        status = normalize_status(cr.get("status"))
        cr_type = normalize_cr_type(cr.get("cr_type"))
        rows.append(
            {
                "release_order": normalize_release_order(cr.get("release_order")),
                "cr_type": CR_TYPE_META[cr_type]["label"],
                "cr_key": sanitize_text(cr.get("cr_key")),
                "created_by": sanitize_text(cr.get("created_by")),
                "description": sanitize_text(cr.get("description")),
                "notes": sanitize_text(cr.get("notes")),
                "status": STATUS_META[status]["label"],
            }
        )

    return {
        "client_name": client_name,
        "project_name": project_name,
        "rows": rows,
    }


def autosize_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 48)


def build_excel_workbook(
    search_text: str = "",
    status_filter: str = "all",
    client_filter: str = "all",
    project_filter: str = "all",
    type_filter: str = "all",
) -> io.BytesIO:
    rows = build_export_rows(
        search_text=search_text,
        status_filter=status_filter,
        client_filter=client_filter,
        project_filter=project_filter,
        type_filter=type_filter,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Lista CR"

    headers = [
        "Cliente",
        "Progetto",
        "Ordine",
        "Tipo CR",
        "Richiesta",
        "Creata da",
        "Descrizione",
        "Note",
        "Stato",
    ]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = cell.font.copy(bold=True)

    for row in rows:
        worksheet.append(
            [
                row["client"],
                row["project"],
                row["release_order"],
                row["cr_type"],
                row["cr_key"],
                row["created_by"],
                row["description"],
                row["notes"],
                row["status"],
            ]
        )

    worksheet.freeze_panes = "A2"
    autosize_worksheet_columns(worksheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_project_excel_workbook(client_id: str, project_id: str) -> io.BytesIO:
    payload = build_project_export_rows(client_id, project_id)
    rows = payload["rows"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CR Progetto"

    worksheet.append(["Cliente", payload["client_name"]])
    worksheet.append(["Progetto", payload["project_name"]])
    worksheet.append(["Generato il", now_it_datetime()])
    worksheet.append([])

    headers = [
        "Ordine",
        "Tipo CR",
        "Richiesta",
        "Creata da",
        "Descrizione",
        "Note",
        "Stato",
    ]
    worksheet.append(headers)

    header_row_index = worksheet.max_row
    for cell in worksheet[header_row_index]:
        cell.font = cell.font.copy(bold=True)

    for row in rows:
        worksheet.append(
            [
                row["release_order"],
                row["cr_type"],
                row["cr_key"],
                row["created_by"],
                row["description"],
                row["notes"],
                row["status"],
            ]
        )

    worksheet.freeze_panes = f"A{header_row_index + 1}"
    autosize_worksheet_columns(worksheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_project_pdf_document(client_id: str, project_id: str) -> io.BytesIO:
    payload = build_project_export_rows(client_id, project_id)
    rows = payload["rows"]

    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"CR_{payload['project_name']}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ProjectTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.HexColor("#0f766e"),
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#374151"),
        leading=14,
    )

    story = [
        Paragraph("SAP CR Manager - Export Progetto", title_style),
        Paragraph(f"<b>Cliente:</b> {payload['client_name']}", meta_style),
        Paragraph(f"<b>Progetto:</b> {payload['project_name']}", meta_style),
        Paragraph(f"<b>Generato il:</b> {now_it_datetime()}", meta_style),
        Spacer(1, 8),
    ]

    table_data = [["Ordine", "Tipo", "Richiesta", "Creata da", "Stato", "Descrizione", "Note"]]
    for row in rows:
        table_data.append(
            [
                row["release_order"],
                row["cr_type"],
                row["cr_key"],
                row["created_by"],
                row["status"],
                row["description"],
                row["notes"],
            ]
        )

    if len(table_data) == 1:
        table_data.append(["-", "-", "Nessuna CR", "-", "-", "-", "-"])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[16 * mm, 25 * mm, 26 * mm, 30 * mm, 24 * mm, 76 * mm, 76 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (0, 0), (4, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f7fbfa"), colors.HexColor("#eef7f5")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9ca3af")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(table)
    document.build(story)
    output.seek(0)
    return output


def build_view_model(
    search_text: str = "",
    status_filter: str = "all",
    client_filter: str = "all",
    project_filter: str = "all",
    type_filter: str = "all",
) -> dict:
    data = load_data()
    normalized_search = sanitize_text(search_text).lower()
    normalized_status = sanitize_text(status_filter).lower() or "all"
    normalized_client = normalize_filter_value(client_filter) or "all"
    normalized_project = normalize_filter_value(project_filter) or "all"
    normalized_type = normalize_filter_value(type_filter) or "all"
    if normalized_status != "all" and normalized_status not in STATUS_META:
        normalized_status = "all"
    if normalized_type != "all" and normalized_type not in CR_TYPE_META:
        normalized_type = "all"

    visible_clients: list[dict] = []
    global_crs: list[dict] = []
    total_projects = 0
    visible_projects = 0
    total_crs = 0
    visible_crs = 0
    filter_options = build_filter_options(data)

    sorted_clients = sorted(data["clients"], key=lambda item: item.get("name", "").lower())
    for client in sorted_clients:
        client_name = sanitize_text(client.get("name"))
        if normalized_client != "all" and client_name.lower() != normalized_client:
            total_projects += len(sorted(client.get("projects", []), key=lambda item: item.get("name", "").lower()))
            total_crs += sum(len(project.get("crs", [])) for project in client.get("projects", []))
            continue

        client_projects: list[dict] = []
        sorted_projects = sorted(client.get("projects", []), key=lambda item: item.get("name", "").lower())
        total_projects += len(sorted_projects)

        for project in sorted_projects:
            project_name = sanitize_text(project.get("name"))
            if normalized_project != "all" and project_name.lower() != normalized_project:
                total_crs += len(project.get("crs", []))
                continue

            decorated_crs = []
            sorted_crs = sort_crs_for_execution(project.get("crs", []))
            total_crs += len(sorted_crs)

            for cr in sorted_crs:
                if not matches_filters(cr, normalized_search, normalized_status, normalized_type):
                    continue

                status = normalize_status(cr.get("status"))
                cr_type = normalize_cr_type(cr.get("cr_type"))
                meta = STATUS_META[status]
                decorated_cr = deepcopy(cr)
                decorated_cr["status_label"] = meta["label"]
                decorated_cr["tone"] = meta["tone"]
                decorated_cr["cr_type"] = cr_type
                decorated_cr["cr_type_label"] = CR_TYPE_META[cr_type]["label"]
                decorated_cr["release_order"] = normalize_release_order(cr.get("release_order"))
                decorated_cr["client_name"] = client_name
                decorated_cr["project_name"] = project_name
                decorated_cr["client_id"] = client["id"]
                decorated_cr["project_id"] = project["id"]
                decorated_cr["created_at"] = format_it_datetime(cr.get("created_at", ""))
                decorated_cr["updated_at"] = format_it_datetime(cr.get("updated_at", ""))
                client_palette = build_client_band_palette(client_name)
                decorated_cr["client_band_bg"] = client_palette["bg"]
                decorated_cr["client_band_border"] = client_palette["border"]
                decorated_cr["client_band_accent"] = client_palette["accent"]
                decorated_crs.append(decorated_cr)
                global_crs.append(decorated_cr)

            kanban_columns = []
            for status_key in KANBAN_ORDER:
                meta = STATUS_META[status_key]
                kanban_columns.append(
                    {
                        "key": status_key,
                        "label": meta["label"],
                        "tone": meta["tone"],
                        "crs": [cr for cr in decorated_crs if cr.get("status") == status_key],
                    }
                )

            visible_projects += 1
            visible_crs += len(decorated_crs)
            project_status_counts = {key: 0 for key in STATUS_META}
            for item in decorated_crs:
                project_status = normalize_status(item.get("status"))
                project_status_counts[project_status] += 1

            client_projects.append(
                {
                    "id": project["id"],
                    "name": project["name"],
                    "crs": decorated_crs,
                    "kanban_columns": kanban_columns,
                    "cr_count": len(project.get("crs", [])),
                    "status_counts": project_status_counts,
                    "next_release_order": next_release_order(project),
                }
            )

        visible_clients.append(
            {
                "id": client["id"],
                "name": client["name"],
                "projects": client_projects,
                "project_count": len(client.get("projects", [])),
                "cr_count": sum(len(project.get("crs", [])) for project in client.get("projects", [])),
            }
        )

    status_counts = {key: 0 for key in STATUS_META}
    for cr in global_crs:
        status = normalize_status(cr.get("status"))
        status_counts[status] += 1

    return {
        "clients": visible_clients,
        "summary": {
            "clients": len(data["clients"]),
            "projects": total_projects,
            "crs": total_crs,
            "visible_clients": len(visible_clients),
            "visible_projects": visible_projects,
            "visible_crs": visible_crs,
            "status_counts": status_counts,
        },
        "filters": {
            "search_text": sanitize_text(search_text),
            "status_filter": normalized_status,
            "client_filter": normalized_client,
            "project_filter": normalized_project,
            "type_filter": normalized_type,
        },
        "filter_options": filter_options,
        "status_meta": STATUS_META,
        "cr_type_meta": CR_TYPE_META,
        "global_kanban_columns": build_global_kanban_columns(global_crs),
        "generated_at": now_it_datetime(),
    }


@app.get("/")
def index():
    view_model = build_view_model(
        search_text=request.args.get("q", ""),
        status_filter=request.args.get("status", "all"),
        client_filter=request.args.get("client", "all"),
        project_filter=request.args.get("project", "all"),
        type_filter=request.args.get("type", "all"),
    )
    return render_template("index.html", **view_model)


@app.get("/export/excel")
def export_excel():
    search_text = request.args.get("q", "")
    status_filter = request.args.get("status", "all")
    client_filter = request.args.get("client", "all")
    project_filter = request.args.get("project", "all")
    type_filter = request.args.get("type", "all")
    workbook_stream = build_excel_workbook(
        search_text=search_text,
        status_filter=status_filter,
        client_filter=client_filter,
        project_filter=project_filter,
        type_filter=type_filter,
    )
    export_project_name = project_filter if sanitize_text(project_filter).lower() != "all" else "TuttiProgetti"
    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=build_project_export_filename(export_project_name, "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/clients/<client_id>/projects/<project_id>/export/excel")
def export_project_excel(client_id: str, project_id: str):
    try:
        workbook_stream = build_project_excel_workbook(client_id, project_id)
        payload = build_project_export_rows(client_id, project_id)
    except ValueError as error:
        flash(str(error), "error")
        return redirect(url_for("index"))

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=build_project_export_filename(payload["project_name"], "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/clients/<client_id>/projects/<project_id>/export/pdf")
def export_project_pdf(client_id: str, project_id: str):
    try:
        pdf_stream = build_project_pdf_document(client_id, project_id)
        payload = build_project_export_rows(client_id, project_id)
    except ValueError as error:
        flash(str(error), "error")
        return redirect(url_for("index"))

    return send_file(
        pdf_stream,
        as_attachment=True,
        download_name=build_project_export_filename(payload["project_name"], "pdf"),
        mimetype="application/pdf",
    )


@app.post("/api/clients/<client_id>/projects/<project_id>/crs/<cr_id>/status")
def update_cr_status_api(client_id: str, project_id: str, cr_id: str):
    payload = request.get_json(silent=True) or {}
    status = normalize_status(payload.get("status", ""))

    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        return jsonify({"ok": False, "message": "Cliente non trovato."}), 404

    project = find_project(client, project_id)
    if project is None:
        return jsonify({"ok": False, "message": "Progetto non trovato."}), 404

    cr = find_cr(project, cr_id)
    if cr is None:
        return jsonify({"ok": False, "message": "CR non trovata."}), 404

    cr["status"] = status
    cr["updated_at"] = now_iso()
    save_data(data)

    return jsonify(
        {
            "ok": True,
            "status": status,
            "status_label": STATUS_META[status]["label"],
            "updated_at": cr["updated_at"],
        }
    )


@app.post("/clients")
def add_client():
    client_name = sanitize_text(request.form.get("client_name"))
    if not client_name:
        flash("Inserisci il nome del cliente.", "error")
        return redirect(url_for("index"))

    data = load_data()
    if any(client.get("name", "").lower() == client_name.lower() for client in data["clients"]):
        flash("Esiste gia un cliente con questo nome.", "error")
        return redirect(url_for("index"))

    data["clients"].append(
        {
            "id": new_id(),
            "name": client_name,
            "projects": [],
        }
    )
    save_data(data)
    flash("Cliente aggiunto.", "success")
    return redirect(url_for("index"))


@app.post("/clients/<client_id>/delete")
def delete_client(client_id: str):
    data = load_data()
    before = len(data["clients"])
    data["clients"] = [client for client in data["clients"] if client.get("id") != client_id]
    if len(data["clients"]) == before:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    save_data(data)
    flash("Cliente eliminato.", "success")
    return redirect(url_for("index"))


@app.post("/clients/<client_id>/projects")
def add_project(client_id: str):
    project_name = sanitize_text(request.form.get("project_name"))
    if not project_name:
        flash("Inserisci il nome del progetto.", "error")
        return redirect(url_for("index"))

    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    if any(project.get("name", "").lower() == project_name.lower() for project in client.get("projects", [])):
        flash("Esiste gia un progetto con questo nome per il cliente selezionato.", "error")
        return redirect(url_for("index"))

    client["projects"].append(
        {
            "id": new_id(),
            "name": project_name,
            "crs": [],
        }
    )
    save_data(data)
    flash("Progetto aggiunto.", "success")
    return redirect(url_for("index"))


@app.post("/clients/<client_id>/projects/<project_id>/delete")
def delete_project(client_id: str, project_id: str):
    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    before = len(client.get("projects", []))
    client["projects"] = [project for project in client.get("projects", []) if project.get("id") != project_id]
    if len(client["projects"]) == before:
        flash("Progetto non trovato.", "error")
        return redirect(url_for("index"))

    save_data(data)
    flash("Progetto eliminato.", "success")
    return redirect(url_for("index"))


@app.post("/clients/<client_id>/projects/<project_id>/crs")
def add_cr(client_id: str, project_id: str):
    cr_key = sanitize_text(request.form.get("cr_key"))
    created_by = sanitize_text(request.form.get("created_by"))
    description = sanitize_text(request.form.get("description"))
    notes = sanitize_text(request.form.get("notes"))
    cr_type = normalize_cr_type(request.form.get("cr_type"))
    release_order = normalize_release_order(request.form.get("release_order"), fallback=0)

    if not cr_key or not created_by or not description:
        flash("Compila Richiesta, utente creatore e descrizione.", "error")
        return redirect(url_for("index"))

    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    project = find_project(client, project_id)
    if project is None:
        flash("Progetto non trovato.", "error")
        return redirect(url_for("index"))

    if any(item.get("cr_key", "").lower() == cr_key.lower() for item in project.get("crs", [])):
        flash("Esiste gia una CR con questa richiesta nel progetto selezionato.", "error")
        return redirect(url_for("index"))

    minimum_release_order = next_release_order(project)
    if release_order < 1:
        release_order = minimum_release_order

    if release_order < minimum_release_order:
        flash(
            f"Ordine CR non valido. Per questo progetto il prossimo ordine disponibile e {minimum_release_order}.",
            "error",
        )
        return redirect(url_for("index"))

    if has_release_order_conflict(project, release_order):
        flash("Ordine CR gia usato nel progetto. Scegli un ordine diverso.", "error")
        return redirect(url_for("index"))

    timestamp = now_iso()
    project["crs"].append(
        {
            "id": new_id(),
            "cr_key": cr_key,
            "created_by": created_by,
            "description": description,
            "notes": notes,
            "status": "development",
            "cr_type": cr_type,
            "release_order": release_order,
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    )
    save_data(data)
    flash("CR aggiunta in stato Sviluppo.", "success")
    return redirect(url_for("index"))


@app.post("/api/clients/<client_id>/projects/<project_id>/crs/clipboard-import")
def import_crs_from_clipboard(client_id: str, project_id: str):
    payload = request.get_json(silent=True) or {}
    raw_entries = payload.get("entries")

    if not isinstance(raw_entries, list) or not raw_entries:
        return jsonify({"ok": False, "message": "Nessuna CR valida da importare."}), 400

    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        return jsonify({"ok": False, "message": "Cliente non trovato."}), 404

    project = find_project(client, project_id)
    if project is None:
        return jsonify({"ok": False, "message": "Progetto non trovato."}), 404

    existing_keys = {sanitize_text(item.get("cr_key", "")).lower() for item in project.get("crs", [])}
    import_keys = set()
    normalized_entries = []

    for index, raw_entry in enumerate(raw_entries, start=1):
        if not isinstance(raw_entry, dict):
            return jsonify({"ok": False, "message": f"Riga {index}: formato non valido."}), 400

        cr_key = sanitize_text(raw_entry.get("cr_key"))
        created_by = sanitize_text(raw_entry.get("created_by"))
        description = sanitize_text(raw_entry.get("description"))
        cr_type = normalize_cr_type(raw_entry.get("cr_type"))

        if not cr_key or not created_by or not description:
            return jsonify({"ok": False, "message": f"Riga {index}: compila Richiesta, creatore e descrizione."}), 400

        lowered_key = cr_key.lower()
        if lowered_key in existing_keys:
            return jsonify({"ok": False, "message": f"Riga {index}: la richiesta {cr_key} esiste gia nel progetto."}), 400

        if lowered_key in import_keys:
            return jsonify({"ok": False, "message": f"Riga {index}: richiesta duplicata nel blocco incollato ({cr_key})."}), 400

        import_keys.add(lowered_key)
        normalized_entries.append(
            {
                "cr_key": cr_key,
                "created_by": created_by,
                "description": description,
                "cr_type": cr_type,
            }
        )

    start_release_order = next_release_order(project)
    timestamp = now_iso()

    total_entries = len(normalized_entries)
    for offset, entry in enumerate(normalized_entries):
        release_order = start_release_order + (total_entries - 1 - offset)
        project.setdefault("crs", []).append(
            {
                "id": new_id(),
                "cr_key": entry["cr_key"],
                "created_by": entry["created_by"],
                "description": entry["description"],
                "notes": "",
                "status": "development",
                "cr_type": entry["cr_type"],
                "release_order": release_order,
                "created_at": timestamp,
                "updated_at": timestamp,
            }
        )

    save_data(data)
    return jsonify(
        {
            "ok": True,
            "imported": len(normalized_entries),
            "start_release_order": start_release_order,
        }
    )


@app.post("/clients/<client_id>/projects/<project_id>/crs/<cr_id>/update")
def update_cr(client_id: str, project_id: str, cr_id: str):
    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    project = find_project(client, project_id)
    if project is None:
        flash("Progetto non trovato.", "error")
        return redirect(url_for("index"))

    cr = find_cr(project, cr_id)
    if cr is None:
        flash("CR non trovata.", "error")
        return redirect(url_for("index"))

    cr_key = sanitize_text(request.form.get("cr_key"))
    created_by = sanitize_text(request.form.get("created_by"))
    description = sanitize_text(request.form.get("description"))
    notes = sanitize_text(request.form.get("notes"))
    cr_type = normalize_cr_type(request.form.get("cr_type"))

    if not cr_key or not created_by or not description:
        flash("Richiesta, utente creatore e descrizione sono obbligatori.", "error")
        return redirect(url_for("index"))

    if any(item.get("id") != cr_id and item.get("cr_key", "").lower() == cr_key.lower() for item in project.get("crs", [])):
        flash("Nel progetto esiste gia un'altra CR con questa richiesta.", "error")
        return redirect(url_for("index"))

    cr["cr_key"] = cr_key
    cr["created_by"] = created_by
    cr["description"] = description
    cr["notes"] = notes
    cr["cr_type"] = cr_type
    cr["updated_at"] = now_iso()
    save_data(data)
    flash("CR aggiornata.", "success")
    return redirect(url_for("index"))


@app.post("/clients/<client_id>/projects/<project_id>/crs/<cr_id>/delete")
def delete_cr(client_id: str, project_id: str, cr_id: str):
    data = load_data()
    client = find_client(data, client_id)
    if client is None:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("index"))

    project = find_project(client, project_id)
    if project is None:
        flash("Progetto non trovato.", "error")
        return redirect(url_for("index"))

    before = len(project.get("crs", []))
    project["crs"] = [cr for cr in project.get("crs", []) if cr.get("id") != cr_id]
    if len(project["crs"]) == before:
        flash("CR non trovata.", "error")
        return redirect(url_for("index"))

    save_data(data)
    flash("CR eliminata.", "success")
    return redirect(url_for("index"))


@app.get("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(debug=APP_DEBUG, use_reloader=APP_DEBUG, host="0.0.0.0", port=APP_PORT)